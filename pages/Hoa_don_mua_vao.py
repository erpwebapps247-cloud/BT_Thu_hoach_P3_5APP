import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from PIL import Image
import pytesseract
from pdf2image import convert_from_bytes
import re
import json

# Import OpenAI (optional)
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

# Đọc API key từ config (nếu có)
try:
    from config import OPENAI_API_KEY as DEFAULT_API_KEY
except ImportError:
    DEFAULT_API_KEY = None

st.set_page_config(
    page_title="Hóa đơn mua vào",
    page_icon="📄",
    layout="wide"
)

st.title("📄 HÓA ĐƠN MUA VÀO")
st.markdown("---")

# Cấu hình tesseract (nếu cần)
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

EXCEL_FILE = "Ket_qua_Hoa_don_mua_vao.xlsx"
SHEET_NAME = "HD_MV"

def fix_vietnamese_accents(text):
    """Sửa lại dấu tiếng Việt bị OCR đọc sai"""
    if not text:
        return text
    
    result = text
    
    # Sửa các từ phổ biến bị OCR đọc sai - sử dụng regex để sửa từ hoàn chỉnh
    # "TON" -> "TÔN" (trong tên công ty)
    result = re.sub(r'\bTON\b', 'TÔN', result, flags=re.IGNORECASE)
    # "THANH" -> "THÀNH" (thành phố, thành công)
    result = re.sub(r'\bTHANH\b', 'THÀNH', result, flags=re.IGNORECASE)
    # "DAT" -> "ĐẠT" (đạt được)
    result = re.sub(r'\bDAT\b', 'ĐẠT', result, flags=re.IGNORECASE)
    # "CONG" -> "CÔNG" (công ty)
    result = re.sub(r'\bCONG\b', 'CÔNG', result, flags=re.IGNORECASE)
    
    # Sửa các trường hợp đặc biệt
    # "THÉP" có thể bị đọc sai thành "THÉP" hoặc "THÉP", giữ nguyên nếu đã có dấu
    # "DONG" có thể là "ĐÔNG" hoặc "ĐỒNG" - tùy ngữ cảnh, nhưng thường trong tên công ty là "ĐÔNG"
    result = re.sub(r'\bDONG\b', 'ĐÔNG', result, flags=re.IGNORECASE)
    
    return result


def extract_invoice_info(image):
    """Trích xuất thông tin từ ảnh hóa đơn sử dụng OCR"""
    try:
        # Sử dụng OCR cơ bản với ngôn ngữ tiếng Việt và tiếng Anh
        text = pytesseract.image_to_string(image, lang='vie+eng')
        return text
    except Exception as e:
        st.error(f"Lỗi khi đọc OCR: {str(e)}")
        return None

def extract_with_openai(text, api_key):
    """Sử dụng OpenAI API để trích xuất thông tin từ text OCR"""
    if not OPENAI_AVAILABLE:
        return None
    
    try:
        client = OpenAI(api_key=api_key)
        
        prompt = f"""Bạn là chuyên gia trích xuất thông tin từ hóa đơn. Hãy phân tích text OCR sau đây và trích xuất thông tin theo định dạng JSON.

Text OCR (có thể có lỗi dấu tiếng Việt do OCR):
{text}

Hãy trích xuất các thông tin sau:
1. SỐ HĐ: Số hóa đơn (ví dụ: 00000788)
2. NGÀY: Ngày hóa đơn (format: DD/MM/YYYY)
3. NỘI DUNG: Danh sách hàng hóa/dịch vụ từ bảng "Tên hàng hóa, dịch vụ". Format mỗi dòng: "STT. Tên hàng hóa" (ví dụ: "1. Polyol Greenfoam GM - 101.1 - WB1")
4. ĐƠN VỊ XUẤT: Tên công ty/đơn vị xuất hóa đơn - QUAN TRỌNG: OCR có thể đọc sai dấu tiếng Việt (ví dụ: "TON" -> "TÔN", "THANH" -> "THÀNH", "DAT" -> "ĐẠT"). Bạn phải TỰ ĐỘNG SỬA LẠI dấu tiếng Việt cho đúng dựa trên ngữ cảnh. Ví dụ: "CÔNG TY TNHH TON THÉP THANH DAT" -> "CÔNG TY TNHH TÔN THÉP THÀNH ĐẠT"
5. GIÁ TRỊ SAU THUẾ: Tổng giá trị sau thuế (chỉ số, không có dấu phẩy hoặc chấm)

Trả về JSON với format:
{{
    "SỐ HĐ": "00000788",
    "NGÀY": "17/01/2026",
    "NỘI DUNG": "1. Polyol Greenfoam GM - 101.1 - WB1\\n2. TẤM NHỰA POLYCARBONATE RỖNG\\n3. Tôn lạnh màu\\n4. Tôn lạnh màu",
    "ĐƠN VỊ XUẤT": "CÔNG TY TNHH TÔN THÉP THÀNH ĐẠT",
    "GIÁ TRỊ SAU THUẾ": "1000000"
}}

LƯU Ý QUAN TRỌNG:
- ĐƠN VỊ XUẤT: OCR thường đọc sai dấu tiếng Việt. Bạn PHẢI tự động sửa lại dựa trên kiến thức tiếng Việt và ngữ cảnh. Ví dụ:
  * "TON" -> "TÔN" (kim loại)
  * "THANH" -> "THÀNH" (thành công, thành phố)
  * "DAT" -> "ĐẠT" (đạt được)
  * "DONG" -> "ĐÔNG" (phía đông)
  * "DONG" -> "ĐỒNG" (tiền, kim loại) - tùy ngữ cảnh
- Luôn sử dụng dấu tiếng Việt CHÍNH XÁC trong tên công ty/đơn vị
- GIÁ TRỊ SAU THUẾ: Chỉ số thuần túy, không có dấu phẩy hoặc chấm

Chỉ trả về JSON, không có text thêm."""
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",  # Sử dụng gpt-4o-mini hoặc gpt-4 nếu có
            messages=[
                {"role": "system", "content": "Bạn là chuyên gia trích xuất thông tin từ hóa đơn. Trả về kết quả dưới dạng JSON chính xác."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=2000
        )
        
        result_text = response.choices[0].message.content.strip()
        
        # Loại bỏ markdown code block nếu có
        if result_text.startswith("```json"):
            result_text = result_text[7:]
        if result_text.startswith("```"):
            result_text = result_text[3:]
        if result_text.endswith("```"):
            result_text = result_text[:-3]
        
        result_text = result_text.strip()
        
        # Parse JSON
        result = json.loads(result_text)
        return result
        
    except json.JSONDecodeError as e:
        st.warning(f"Không thể parse JSON từ OpenAI: {str(e)}")
        return None
    except Exception as e:
        st.error(f"Lỗi khi gọi OpenAI API: {str(e)}")
        return None

def process_extracted_text(extracted_text, use_openai, api_key):
    """Xử lý text đã trích xuất bằng OCR, có thể dùng OpenAI nếu được bật"""
    if not extracted_text:
        return None
    
    # Sử dụng OpenAI nếu được bật và có API key
    if use_openai and api_key and OPENAI_AVAILABLE:
        with st.spinner("🤖 Đang sử dụng OpenAI để trích xuất thông tin..."):
            openai_data = extract_with_openai(extracted_text, api_key)
            if openai_data:
                st.success("✅ Đã sử dụng OpenAI để trích xuất thông tin")
                return openai_data
            else:
                # Fallback về phương pháp cũ
                st.info("ℹ️ Sử dụng phương pháp OCR thông thường")
                return parse_invoice_text(extracted_text)
    else:
        return parse_invoice_text(extracted_text)

def parse_invoice_text(text):
    """Phân tích text OCR để trích xuất thông tin hóa đơn"""
    info = {
        'SỐ HĐ': '',
        'NGÀY': '',
        'NỘI DUNG': '',
        'ĐƠN VỊ XUẤT': '',
        'GIÁ TRỊ SAU THUẾ': ''
    }
    
    if not text:
        return info
    
    # Tìm số hóa đơn - cải thiện pattern để bắt được "SỐ (No.): 00000788"
    # Tìm số có nhiều số 0 đầu (như 00000788) - có thể đứng sau nhiều từ khóa
    
    # Bước 1: Thử các pattern cụ thể trước
    so_hd_patterns = [
        r'(?:SỐ|SO|Số)\s*\(?\s*No\.?\s*\)?\s*[:]?\s*(\d{4,})',  # "SỐ (No.): 00000788"
        r'No\.\s*[:]?\s*(\d{4,})',  # "No.: 00000788"
        r'(?:SỐ|SO|Số)\s*[:]?\s*(\d{4,})',  # "SỐ: 00000788"
        r'(?:Số|SO|Số HĐ|HĐ số|HD)\s*[:]?\s*(\d{4,})',  # Các biến thể
        r'(?:Invoice|INVOICE)\s*No\.?\s*[:]?\s*(\d{4,})',  # "Invoice No: 00000788"
    ]
    
    for pattern in so_hd_patterns:
        matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
        for match in matches:
            so_hd = match.group(1).strip()
            # Chỉ nhận số thuần túy, bỏ qua chữ cái
            if so_hd and so_hd.isdigit() and len(so_hd) >= 4:
                info['SỐ HĐ'] = so_hd
                break
        if info['SỐ HĐ']:
            break
    
    # Bước 2: Nếu chưa tìm thấy, tìm số dạng 0000xxxx (có nhiều số 0 đầu) gần từ khóa
    if not info['SỐ HĐ']:
        # Tìm số có ít nhất 6 chữ số, có thể bắt đầu bằng số 0
        long_number_patterns = [
            r'\b0{3,}\d{4,}\b',  # Số bắt đầu bằng ít nhất 3 số 0
            r'\b\d{6,}\b',  # Bất kỳ số nào có >= 6 chữ số
        ]
        
        for pattern in long_number_patterns:
            matches = re.finditer(pattern, text)
            for match in matches:
                number = match.group(0)
                if not number.isdigit():
                    continue
                
                # Kiểm tra context xung quanh (30 ký tự trước, 10 ký tự sau)
                start_pos = max(0, match.start() - 30)
                end_pos = min(len(text), match.end() + 10)
                context = text[start_pos:end_pos]
                
                # Kiểm tra xem có từ khóa liên quan không
                if re.search(r'(?:SỐ|SO|Số|No\.|Invoice|HĐ|HD)', context, re.IGNORECASE):
                    if len(number) >= 6:
                        info['SỐ HĐ'] = number
                        break
            if info['SỐ HĐ']:
                break
    
    # Bước 3: Nếu vẫn chưa tìm thấy, tìm số đầu tiên có >= 6 chữ số sau từ khóa "SỐ" hoặc "No"
    if not info['SỐ HĐ']:
        # Tìm vị trí của từ khóa
        keyword_match = re.search(r'(?:SỐ|SO|Số|No\.|Invoice)', text, re.IGNORECASE)
        if keyword_match:
            # Tìm số sau từ khóa (trong vòng 50 ký tự)
            search_text = text[keyword_match.end():keyword_match.end() + 50]
            number_match = re.search(r'(\d{6,})', search_text)
            if number_match:
                number = number_match.group(1)
                if number.isdigit():
                    info['SỐ HĐ'] = number
    
    # Tìm ngày
    date_patterns = [
        r'(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})',
        r'(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{2})',
        r'Ngày[\s:]*(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})',
        r'Date[\s:]*(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})'
    ]
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            day, month, year = match.groups()
            if len(year) == 2:
                year = '20' + year
            info['NGÀY'] = f"{day}/{month}/{year}"
            break
    
    # Tìm đơn vị xuất - sử dụng pattern rộng để giữ tất cả ký tự tiếng Việt
    don_vi_patterns = [
        r'(?:Đơn vị|Công ty|CÔNG TY|ĐƠN VỊ|Company)[\s:]*([^\n]+?)(?=\n|$)',
        r'(?:Bán bởi|Seller|Người bán)[\s:]*([^\n]+?)(?=\n|$)'
    ]
    for pattern in don_vi_patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            don_vi = match.group(1).strip()
            # Loại bỏ các ký tự đặc biệt không cần thiết ở cuối nhưng giữ dấu tiếng Việt
            don_vi = re.sub(r'[\s\-\.]+$', '', don_vi)
            # Chỉ lấy nếu có ít nhất 3 ký tự và có chữ cái
            if len(don_vi) >= 3 and re.search(r'[A-Za-zÀ-ỹ]', don_vi):
                # Sửa lại dấu tiếng Việt bị OCR đọc sai
                don_vi = fix_vietnamese_accents(don_vi)
                info['ĐƠN VỊ XUẤT'] = don_vi
                break
    
    # Tìm nội dung/nhóm hàng - trích xuất từ bảng "Tên hàng hóa, dịch vụ"
    noi_dung_items = []
    
    # Đơn giản hóa: Tìm tất cả các dòng có format "số. text" hoặc "số text" trong toàn bộ text
    # Pattern đơn giản và linh hoạt hơn
    row_patterns = [
        r'(?:^|\n)\s*(\d{1,2})\.\s+([^\n]{3,150})',  # "1. text" - pattern chính
        r'(?:^|\n)\s*(\d{1,2})\s+([^\n]{3,150})',  # "1 text" - dự phòng
    ]
    
    all_found_rows = []
    for pattern in row_patterns:
        matches = re.finditer(pattern, text, re.MULTILINE)
        for match in matches:
            row_num = match.group(1)
            item_name = match.group(2).strip()
            
            # Loại bỏ khoảng trắng thừa
            item_name = item_name.strip()
            
            # Loại bỏ các ký tự không cần thiết ở cuối
            item_name = re.sub(r'[\s\-\.]+$', '', item_name)
            
            # Kiểm tra hợp lệ: có chứa chữ cái, độ dài >= 3
            if (item_name and 
                len(item_name) >= 3 and 
                re.search(r'[A-Za-zÀ-ỹ]', item_name) and
                not item_name.replace('-', '').replace('.', '').replace(' ', '').isdigit() and
                not re.match(r'^(?:STT|No|SỐ|Tổng|Total|Ngày|Date|Đơn|vị|Tên hàng|Name)', item_name, re.IGNORECASE)):
                
                all_found_rows.append((int(row_num), item_name))
    
    # Loại bỏ trùng lặp và sắp xếp
    seen = set()
    unique_rows = []
    for row_num, item_name in sorted(all_found_rows, key=lambda x: x[0]):
        # Chỉ lấy số thứ tự từ 1-10
        if row_num < 1 or row_num > 10:
            continue
        key = (row_num, item_name.lower())
        if key not in seen:
            seen.add(key)
            unique_rows.append((row_num, item_name))
    
    # Lấy các dòng liên tiếp bắt đầu từ 1 (1, 2, 3, 4...)
    if unique_rows:
        consecutive_items = []
        expected_num = 1
        
        for row_num, item_name in unique_rows:
            if row_num == expected_num:
                formatted_item = f"{row_num}. {item_name}"
                consecutive_items.append(formatted_item)
                expected_num += 1
            elif row_num > expected_num:
                # Nếu đã có ít nhất 2 dòng, dừng lại
                if len(consecutive_items) >= 2:
                    break
        
        # Nếu tìm thấy ít nhất 2 dòng liên tiếp, sử dụng kết quả
        if len(consecutive_items) >= 2:
            info['NỘI DUNG'] = '\n'.join(consecutive_items)
        elif len(unique_rows) >= 2:
            # Nếu không có nhóm liên tiếp, lấy các dòng từ 1-4
            filtered_rows = [(r, n) for r, n in unique_rows if 1 <= r <= 4]
            if filtered_rows:
                for row_num, item_name in filtered_rows:
                    formatted_item = f"{row_num}. {item_name}"
                    noi_dung_items.append(formatted_item)
                if noi_dung_items:
                    info['NỘI DUNG'] = '\n'.join(noi_dung_items)
    
    # Fallback: Nếu vẫn chưa có, thử tìm các dòng đơn giản hơn
    if not info['NỘI DUNG']:
        # Tìm các dòng có số ở đầu (1-10) và text sau đó (ít nhất 5 ký tự)
        simple_pattern = r'\n\s*([1-9]|10)[\.\s]+([A-Za-zÀ-ỹ][^\n]{4,100})'
        simple_matches = re.findall(simple_pattern, text, re.MULTILINE | re.IGNORECASE)
        
        if simple_matches and len(simple_matches) >= 2:
            fallback_items = []
            for row_num_str, item_name in simple_matches:
                row_num = int(row_num_str)
                item_name = item_name.strip()
                if (item_name and 
                    len(item_name) >= 3 and 
                    re.search(r'[A-Za-zÀ-ỹ]', item_name) and
                    not re.match(r'^(?:STT|No|SỐ|Tổng|Total)', item_name, re.IGNORECASE)):
                    formatted_item = f"{row_num}. {item_name}"
                    fallback_items.append((row_num, formatted_item))
            
            if fallback_items:
                # Sắp xếp và lấy các dòng liên tiếp từ 1
                fallback_items.sort(key=lambda x: x[0])
                final_items = []
                expected = 1
                for num, item in fallback_items:
                    if num == expected:
                        final_items.append(item)
                        expected += 1
                    elif num > expected and len(final_items) >= 2:
                        break
                
                if len(final_items) >= 2:
                    info['NỘI DUNG'] = '\n'.join(final_items)
                elif len(fallback_items) >= 2:
                    # Lấy 4 dòng đầu tiên
                    info['NỘI DUNG'] = '\n'.join([item for _, item in fallback_items[:4]])
    
    # Tìm giá trị sau thuế
    gia_tri_patterns = [
        r'(?:Tổng|Total|Thành tiền|Sau thuế|SAU THUẾ|Số tiền)[\s:]*[\d.,]*\s*([\d.,]+)',
        r'([\d.,]+)[\s]*VND',
        r'([\d.,]+)[\s]*đ'
    ]
    for pattern in gia_tri_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            value = match.group(1).replace(',', '').replace('.', '')
            if value.isdigit():
                info['GIÁ TRỊ SAU THUẾ'] = value
                break
    
    return info

def load_excel_data():
    """Đọc dữ liệu từ file Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
        else:
            ws = wb[SHEET_NAME]
        
        # Lấy dữ liệu
        data = []
        headers = ['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ']
        
        # Kiểm tra xem đã có header chưa
        if ws.max_row == 0 or ws.cell(1, 1).value is None:
            ws.append(headers)
        
        # Đọc dữ liệu từ hàng 2 trở đi
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                data.append(row)
        
        wb.save(EXCEL_FILE)
        return pd.DataFrame(data, columns=headers) if data else pd.DataFrame(columns=headers)
    except FileNotFoundError:
        # Tạo file mới nếu chưa tồn tại
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        headers = ['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ']
        ws.append(headers)
        wb.save(EXCEL_FILE)
        return pd.DataFrame(columns=headers)
    except Exception as e:
        st.error(f"Lỗi khi đọc file Excel: {str(e)}")
        return pd.DataFrame(columns=['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ'])

def save_to_excel(new_data):
    """Ghi dữ liệu mới vào file Excel với định dạng font tiếng Việt và độ rộng cột"""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            headers = ['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ']
            ws.append(headers)
        else:
            ws = wb[SHEET_NAME]
        
        # Kiểm tra xem đã có header chưa
        if ws.max_row == 0 or ws.cell(1, 1).value is None:
            headers = ['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ']
            ws.append(headers)
        
        # Định dạng header: font tiếng Việt, đậm, nền xanh
        headers = ['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Thêm dữ liệu mới
        ws.append([
            new_data.get('SỐ HĐ', ''),
            new_data.get('NGÀY', ''),
            new_data.get('NỘI DUNG', ''),
            new_data.get('ĐƠN VỊ XUẤT', ''),
            new_data.get('GIÁ TRỊ SAU THUẾ', '')
        ])
        
        # Định dạng dữ liệu: font tiếng Việt, wrap text cho các cột dài
        data_font = Font(name="Arial", size=10)
        column_widths = {
            'A': 15,  # SỐ HĐ
            'B': 15,  # NGÀY
            'C': 60,  # NỘI DUNG
            'D': 50,  # ĐƠN VỊ XUẤT
            'E': 20   # GIÁ TRỊ SAU THUẾ
        }
        
        # Điều chỉnh độ rộng cột
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Định dạng dữ liệu cho hàng mới
        new_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(new_row, col_idx)
            cell.font = data_font
            # Wrap text cho các cột nội dung dài
            if col_idx in [3, 4]:  # NỘI DUNG, ĐƠN VỊ XUẤT
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        st.error(f"Lỗi khi ghi file Excel: {str(e)}")
        import traceback
        st.error(f"Chi tiết lỗi: {traceback.format_exc()}")
        return False

# UI chính
tab1, tab2 = st.tabs(["📤 Nhập hóa đơn mới", "📋 Danh sách hóa đơn"])

with tab1:
    st.header("Nhập hóa đơn từ file PDF hoặc ảnh")
    
    # Cấu hình OpenAI (nếu có)
    with st.expander("🔧 Cấu hình nâng cao (OpenAI API)", expanded=False):
        # Mặc định bật OpenAI nếu có API key
        default_use_openai = True if DEFAULT_API_KEY and OPENAI_AVAILABLE else False
        use_openai = st.checkbox("Sử dụng OpenAI API để trích xuất thông tin chính xác hơn", value=default_use_openai)
        if use_openai:
            if not OPENAI_AVAILABLE:
                st.error("⚠️ Thư viện OpenAI chưa được cài đặt. Vui lòng chạy: pip install openai")
                api_key = None
            else:
                api_key = st.text_input(
                    "OpenAI API Key",
                    type="password",
                    help="Nhập API key của bạn từ https://platform.openai.com/api-keys",
                    value=st.session_state.get('openai_api_key', DEFAULT_API_KEY or '')
                )
                if api_key:
                    st.session_state['openai_api_key'] = api_key
                    st.success("✅ API Key đã được lưu")
        else:
            api_key = None
            use_openai = False
    
    # Khởi tạo biến nếu chưa có
    if 'use_openai' not in locals():
        use_openai = False
    if 'api_key' not in locals():
        api_key = st.session_state.get('openai_api_key', DEFAULT_API_KEY)
    
    uploaded_file = st.file_uploader(
        "Chọn file PDF hoặc ảnh hóa đơn",
        type=['pdf', 'png', 'jpg', 'jpeg'],
        help="Hỗ trợ file PDF hoặc ảnh (PNG, JPG, JPEG)"
    )
    
    if uploaded_file is not None:
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Xem trước file")
            if uploaded_file.type == 'application/pdf':
                st.info("📄 File PDF đã được tải lên")
                # Chuyển đổi PDF thành ảnh
                try:
                    pdf_bytes = uploaded_file.read()
                    images = convert_from_bytes(pdf_bytes, dpi=200)
                    if images:
                        st.image(images[0], caption="Trang đầu của PDF", use_container_width=True)
                        extracted_text = extract_invoice_info(images[0])
                        invoice_data = process_extracted_text(extracted_text, use_openai, api_key)
                    else:
                        st.error("Không thể đọc file PDF")
                        invoice_data = None
                except Exception as e:
                    st.error(f"Lỗi khi xử lý PDF: {str(e)}")
                    invoice_data = None
            else:
                # Xử lý ảnh
                image = Image.open(uploaded_file)
                st.image(image, caption="Ảnh hóa đơn", use_container_width=True)
                extracted_text = extract_invoice_info(image)
                invoice_data = process_extracted_text(extracted_text, use_openai, api_key)
        
        with col2:
            st.subheader("Thông tin trích xuất")
            
            if invoice_data:
                # Hiển thị text OCR (có thể ẩn)
                with st.expander("📝 Text OCR đã đọc"):
                    st.text_area("", extracted_text, height=200, disabled=True)
                
                # Form chỉnh sửa thông tin
                st.markdown("**Vui lòng kiểm tra và chỉnh sửa thông tin:**")
                
                so_hd = st.text_input("Số HĐ", value=invoice_data['SỐ HĐ'])
                ngay = st.text_input("Ngày", value=invoice_data['NGÀY'])
                noi_dung = st.text_area("Nội dung", value=invoice_data['NỘI DUNG'])
                don_vi = st.text_input("Đơn vị xuất", value=invoice_data['ĐƠN VỊ XUẤT'])
                gia_tri = st.text_input("Giá trị sau thuế", value=invoice_data['GIÁ TRỊ SAU THUẾ'])
                
                if st.button("💾 Lưu hóa đơn vào Excel", type="primary"):
                    final_data = {
                        'SỐ HĐ': so_hd,
                        'NGÀY': ngay,
                        'NỘI DUNG': noi_dung,
                        'ĐƠN VỊ XUẤT': don_vi,
                        'GIÁ TRỊ SAU THUẾ': gia_tri if gia_tri else ''
                    }
                    
                    if save_to_excel(final_data):
                        st.success("✅ Đã lưu hóa đơn thành công!")
                        st.balloons()
                    else:
                        st.error("❌ Lỗi khi lưu hóa đơn")
            else:
                st.warning("Không thể trích xuất thông tin từ file")

with tab2:
    st.header("Danh sách hóa đơn đã lưu")
    
    df = load_excel_data()
    
    if not df.empty:
        st.dataframe(df, use_container_width=True)
        
        # Thống kê
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Tổng số hóa đơn", len(df))
        with col2:
            if 'GIÁ TRỊ SAU THUẾ' in df.columns:
                try:
                    # Loại bỏ dấu cách, phẩy và chuyển sang số
                    total = pd.to_numeric(
                        df['GIÁ TRỊ SAU THUẾ'].astype(str).str.replace(' ', '').str.replace(',', '').replace('', '0'),
                        errors='coerce'
                    ).sum()
                    # Format với dấu cách ngàn
                    total_formatted = f"{int(total):,}".replace(',', ' ') if not pd.isna(total) else "0"
                    st.metric("Tổng giá trị", total_formatted)
                except Exception:
                    st.metric("Tổng giá trị", "N/A")
        with col3:
            if st.button("🔄 Làm mới dữ liệu"):
                st.rerun()
    else:
        st.info("Chưa có hóa đơn nào được lưu. Vui lòng nhập hóa đơn mới ở tab 'Nhập hóa đơn mới'")

st.markdown("---")
st.markdown("**File Excel:** `Ket_qua_Hoa_don_mua_vao.xlsx` | **Sheet:** `HD_MV`")
