import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook

st.set_page_config(
    page_title="Kết quả kinh doanh",
    page_icon="📊",
    layout="wide"
)

st.title("📊 KẾT QUẢ KINH DOANH")
st.markdown("---")

# Đường dẫn file Excel
EXCEL_FILE_MUA_VAO = "Ket_qua_Hoa_don_mua_vao.xlsx"
SHEET_NAME_MUA_VAO = "HD_MV"
EXCEL_FILE_BAN_RA = "Ket_qua_Hoa_don_ban_ra.xlsx"
SHEET_NAME_BAN_RA = "HD_BR"

def load_excel_data_mua_vao():
    """Đọc dữ liệu từ file Excel hóa đơn mua vào"""
    try:
        wb = load_workbook(EXCEL_FILE_MUA_VAO)
        if SHEET_NAME_MUA_VAO not in wb.sheetnames:
            return pd.DataFrame(columns=['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ'])
        else:
            ws = wb[SHEET_NAME_MUA_VAO]
            data = []
            headers = ['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ']
            
            # Đọc từ hàng 2 trở đi
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    data.append(row)
            
            return pd.DataFrame(data, columns=headers) if data else pd.DataFrame(columns=headers)
    except FileNotFoundError:
        return pd.DataFrame(columns=['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ'])
    except Exception as e:
        st.error(f"Lỗi khi đọc file Excel mua vào: {str(e)}")
        return pd.DataFrame(columns=['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ XUẤT', 'GIÁ TRỊ SAU THUẾ'])

def load_excel_data_ban_ra():
    """Đọc dữ liệu từ file Excel hóa đơn bán ra"""
    try:
        wb = load_workbook(EXCEL_FILE_BAN_RA)
        if SHEET_NAME_BAN_RA not in wb.sheetnames:
            return pd.DataFrame(columns=['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ NHẬN', 'GIÁ TRỊ SAU THUẾ'])
        else:
            ws = wb[SHEET_NAME_BAN_RA]
            data = []
            headers = ['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ NHẬN', 'GIÁ TRỊ SAU THUẾ']
            
            # Đọc từ hàng 2 trở đi
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    data.append(row)
            
            return pd.DataFrame(data, columns=headers) if data else pd.DataFrame(columns=headers)
    except FileNotFoundError:
        return pd.DataFrame(columns=['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ NHẬN', 'GIÁ TRỊ SAU THUẾ'])
    except Exception as e:
        st.error(f"Lỗi khi đọc file Excel bán ra: {str(e)}")
        return pd.DataFrame(columns=['SỐ HĐ', 'NGÀY', 'NỘI DUNG', 'ĐƠN VỊ NHẬN', 'GIÁ TRỊ SAU THUẾ'])

def calculate_total_value(df, column_name='GIÁ TRỊ SAU THUẾ'):
    """Tính tổng giá trị từ cột trong DataFrame"""
    if df.empty or column_name not in df.columns:
        return 0
    
    try:
        # Loại bỏ dấu cách, phẩy và chuyển sang số
        total = pd.to_numeric(
            df[column_name].astype(str).str.replace(' ', '').str.replace(',', '').str.replace('.', '').replace('', '0'),
            errors='coerce'
        ).sum()
        return total if not pd.isna(total) else 0
    except Exception:
        return 0

def format_number(num):
    """Định dạng số với dấu cách ngàn"""
    try:
        return f"{int(num):,}".replace(',', ' ')
    except:
        return "0"

# Load dữ liệu
df_mua_vao = load_excel_data_mua_vao()
df_ban_ra = load_excel_data_ban_ra()

# Tính toán các chỉ tiêu
tong_gia_tri_mua_vao = calculate_total_value(df_mua_vao, 'GIÁ TRỊ SAU THUẾ')
tong_gia_tri_ban_ra = calculate_total_value(df_ban_ra, 'GIÁ TRỊ SAU THUẾ')
thue_vat_phai_nop = tong_gia_tri_ban_ra - tong_gia_tri_mua_vao
thue_tndn_phai_nop = 0.22 * thue_vat_phai_nop if thue_vat_phai_nop > 0 else 0

# Hiển thị thống kê
st.markdown("### **📈 CÁC CHỈ TIÊU KINH DOANH**")

col1, col2, col3, col4 = st.columns(4)

with col1:
    st.metric(
        label="💰 Tổng giá trị mua vào",
        value=format_number(tong_gia_tri_mua_vao) + " đ",
        help=f"Số hóa đơn: {len(df_mua_vao)}"
    )

with col2:
    st.metric(
        label="💵 Tổng giá trị bán ra",
        value=format_number(tong_gia_tri_ban_ra) + " đ",
        help=f"Số hóa đơn: {len(df_ban_ra)}"
    )

with col3:
    delta_vat = thue_vat_phai_nop
    st.metric(
        label="🧾 Thuế VAT phải nộp",
        value=format_number(thue_vat_phai_nop) + " đ",
        delta=f"{format_number(delta_vat)} đ",
        delta_color="normal" if thue_vat_phai_nop >= 0 else "inverse",
        help="Công thức: Tổng bán ra - Tổng mua vào"
    )

with col4:
    st.metric(
        label="📋 Thuế TNDN phải nộp",
        value=format_number(thue_tndn_phai_nop) + " đ",
        help="Công thức: 22% × Thuế VAT phải nộp"
    )

st.markdown("---")

# Bảng tổng hợp chi tiết
st.markdown("### **📊 BẢNG TỔNG HỢP CHI TIẾT**")

col_left, col_right = st.columns(2)

with col_left:
    st.markdown("#### **📥 Hóa đơn mua vào**")
    if not df_mua_vao.empty:
        st.dataframe(df_mua_vao, use_container_width=True, height=300)
        st.info(f"📊 Tổng số hóa đơn: **{len(df_mua_vao)}** | Tổng giá trị: **{format_number(tong_gia_tri_mua_vao)} đ**")
    else:
        st.info("ℹ️ Chưa có dữ liệu hóa đơn mua vào")

with col_right:
    st.markdown("#### **📤 Hóa đơn bán ra**")
    if not df_ban_ra.empty:
        st.dataframe(df_ban_ra, use_container_width=True, height=300)
        st.info(f"📊 Tổng số hóa đơn: **{len(df_ban_ra)}** | Tổng giá trị: **{format_number(tong_gia_tri_ban_ra)} đ**")
    else:
        st.info("ℹ️ Chưa có dữ liệu hóa đơn bán ra")

st.markdown("---")

# Bảng kết quả tính toán
st.markdown("### **🧮 BẢNG KẾT QUẢ TÍNH TOÁN**")

result_data = {
    'Chỉ tiêu': [
        'Tổng giá trị mua vào',
        'Tổng giá trị bán ra',
        'Thuế VAT phải nộp (HD_BR - HD_MV)',
        'Thuế TNDN phải nộp (22% × VAT)'
    ],
    'Giá trị': [
        f"{format_number(tong_gia_tri_mua_vao)} đ",
        f"{format_number(tong_gia_tri_ban_ra)} đ",
        f"{format_number(thue_vat_phai_nop)} đ",
        f"{format_number(thue_tndn_phai_nop)} đ"
    ],
    'Số tiền (số)': [
        tong_gia_tri_mua_vao,
        tong_gia_tri_ban_ra,
        thue_vat_phai_nop,
        thue_tndn_phai_nop
    ]
}

df_result = pd.DataFrame(result_data)
st.dataframe(df_result[['Chỉ tiêu', 'Giá trị']], use_container_width=True, hide_index=True)

# Công thức tính toán
with st.expander("📐 Chi tiết công thức tính toán", expanded=False):
    st.markdown("""
    **Công thức tính toán:**
    
    1. **Tổng giá trị mua vào** = Tổng cột "GIÁ TRỊ SAU THUẾ" từ file `Ket_qua_Hoa_don_mua_vao.xlsx`, sheet `HD_MV`
    
    2. **Tổng giá trị bán ra** = Tổng cột "GIÁ TRỊ SAU THUẾ" từ file `Ket_qua_Hoa_don_ban_ra.xlsx`, sheet `HD_BR`
    
    3. **Thuế VAT phải nộp** = Tổng giá trị bán ra - Tổng giá trị mua vào
       - Công thức: `HD_BR - HD_MV`
    
    4. **Thuế TNDN phải nộp** = 22% × Thuế VAT phải nộp
       - Công thức: `22% × (HD_BR - HD_MV)`
       - Lưu ý: Chỉ tính khi Thuế VAT phải nộp > 0
    """)

# Nút làm mới
if st.button("🔄 Làm mới dữ liệu", type="primary"):
    st.rerun()

st.markdown("---")
st.markdown("**📁 Dữ liệu nguồn:**")
st.markdown(f"- **Hóa đơn mua vào:** `{EXCEL_FILE_MUA_VAO}` | Sheet: `{SHEET_NAME_MUA_VAO}`")
st.markdown(f"- **Hóa đơn bán ra:** `{EXCEL_FILE_BAN_RA}` | Sheet: `{SHEET_NAME_BAN_RA}`")
