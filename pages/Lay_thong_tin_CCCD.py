import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract
import re
from datetime import datetime
import json

# Import OpenAI (optional)
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

# Đọc API key từ config (nếu có)
try:
    from config import OPENAI_API_KEY as DEFAULT_API_KEY
except ImportError:
    DEFAULT_API_KEY = None

st.set_page_config(
    page_title="Lấy thông tin CCCD",
    page_icon="🆔",
    layout="wide"
)

st.title("🆔 LẤY THÔNG TIN NHÂN VIÊN TỪ CCCD")
st.markdown("---")

# Cấu hình tesseract (nếu cần)
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

EXCEL_FILE = "Ket_qua_CCCD.xlsx"

def extract_text_with_ocr(image):
    """Trích xuất text từ ảnh sử dụng OCR cơ bản"""
    try:
        # Sử dụng OCR cơ bản với ngôn ngữ tiếng Việt và tiếng Anh
        text = pytesseract.image_to_string(image, lang='vie+eng')
        return text
    except Exception as e:
        st.error(f"Lỗi khi đọc OCR: {str(e)}")
        return ""

def extract_cccd_with_openai(text_front, text_back, api_key):
    """Sử dụng OpenAI API để trích xuất thông tin từ text OCR CCCD"""
    if not OPENAI_AVAILABLE:
        return None
    
    try:
        client = OpenAI(api_key=api_key)
        
        full_text = f"MẶT TRƯỚC:\n{text_front}\n\nMẶT SAU:\n{text_back}"
        
        prompt = f"""Bạn là chuyên gia trích xuất thông tin từ CCCD (Căn cước công dân) Việt Nam. Hãy phân tích text OCR sau đây và trích xuất thông tin theo định dạng JSON.

Text OCR từ CCCD:
{full_text}

Hãy trích xuất các thông tin sau (PHẢI GIỮ NGUYÊN dấu tiếng Việt và ĐỌC CHÍNH XÁC số):
1. Số CCCD: Số căn cước công dân (12 chữ số) - thường ở định dạng "Số / No.: 080188012880" hoặc tương tự
2. Họ và tên: Họ và tên đầy đủ (GIỮ NGUYÊN dấu tiếng Việt)
3. Ngày sinh: Format DD/MM/YYYY - ĐỌC CHÍNH XÁC từng số, đặc biệt là năm (ví dụ: 01/01/1988 không phải 01/01/1980)
4. Giới tính: Nam hoặc Nữ
5. Quốc tịch: Thường là "Việt Nam" hoặc "Vietnam"
6. Quê quán: Địa chỉ quê quán (GIỮ NGUYÊN dấu tiếng Việt) - LƯU Ý: Giá trị có thể nằm ở DÒNG DƯỚI sau từ khóa "Quê quán / Place of origin:" và có thể trải dài nhiều dòng. GHÉP TẤT CẢ các dòng lại thành một địa chỉ đầy đủ.
7. Nơi thường trú: Địa chỉ thường trú (GIỮ NGUYÊN dấu tiếng Việt) - LƯU Ý: Giá trị có thể BẮT ĐẦU CÙNG DÒNG với từ khóa (sau dấu :) và TIẾP TỤC ở các dòng dưới. GHÉP TẤT CẢ các dòng lại thành một địa chỉ đầy đủ (ví dụ: "637/10/33/30P Hà Huy Giáp, KP2, Thạnh Xuân, Q12, TP. HCM")
8. Ngày cấp: Format DD/MM/YYYY - ĐỌC CHÍNH XÁC từng số
9. Nơi cấp: Tên cơ quan cấp (GIỮ NGUYÊN dấu tiếng Việt)

LƯU Ý QUAN TRỌNG:
- CÁC THÔNG TIN TRÊN CCCD CÓ THỂ KHÔNG THẲNG HÀNG: Tìm từ khóa (ví dụ: "Ngày sinh / Date of birth:") rồi tìm giá trị trong PHẠM VI RỘNG quanh đó, không chỉ trên cùng một dòng.
- Ví dụ: Nếu thấy "Ngày sinh / Date of birth:" nhưng ngày tháng năm ở dòng khác hoặc bị lệch, vẫn phải trích xuất đúng.
- OCR có thể đọc sai dấu tiếng Việt hoặc số. Bạn PHẢI tự động sửa lại dấu và số cho đúng dựa trên ngữ cảnh và kiến thức tiếng Việt/định dạng CCCD.
- Ví dụ sửa dấu: "TON" -> "TÔN", "THANH" -> "THÀNH", "DAT" -> "ĐẠT", "CONG" -> "CÔNG", "DONG" -> "ĐÔNG"
- NGÀY SINH: Đọc CHÍNH XÁC từng chữ số. Nếu thấy "01/01/1988" thì phải là "01/01/1988", KHÔNG phải "01/01/1980" hay "01/01/1990". Kiểm tra kỹ số cuối cùng của năm (ví dụ: 1988 có số 8 cuối, không phải 0).
- Đảm bảo tất cả thông tin địa chỉ, tên đều có dấu tiếng Việt chính xác
- Số CCCD phải là 12 chữ số và chính xác, tìm sau "Số / No.:"

Trả về JSON với format:
{{
    "Số CCCD": "001234567890",
    "Họ và tên": "NGUYỄN VĂN A",
    "Ngày sinh": "01/01/1990",
    "Giới tính": "Nam",
    "Quốc tịch": "Việt Nam",
    "Quê quán": "Xã ABC, Huyện XYZ, Tỉnh DEF",
    "Nơi thường trú": "Số 123 Đường ABC, Phường XYZ, Thành phố DEF",
    "Ngày cấp": "01/01/2020",
    "Nơi cấp": "CỤC CẢNH SÁT ĐKQL CƯ TRÚ VÀ DLQG VỀ DÂN CƯ"
}}

Chỉ trả về JSON, không có text thêm."""
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Bạn là chuyên gia trích xuất thông tin từ CCCD Việt Nam. Trả về kết quả dưới dạng JSON chính xác với dấu tiếng Việt đúng."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=2000
        )
        
        result_text = response.choices[0].message.content.strip()
        
        # Loại bỏ markdown code block nếu có
        if result_text.startswith("```json"):
            result_text = result_text[7:]
        if result_text.startswith("```"):
            result_text = result_text[3:]
        if result_text.endswith("```"):
            result_text = result_text[:-3]
        
        result_text = result_text.strip()
        
        # Parse JSON
        result = json.loads(result_text)
        return result
        
    except json.JSONDecodeError as e:
        st.warning(f"Không thể parse JSON từ OpenAI: {str(e)}")
        return None
    except Exception as e:
        st.error(f"Lỗi khi gọi OpenAI API: {str(e)}")
        return None

def process_cccd_extraction(image_front, image_back, use_openai, api_key):
    """Xử lý trích xuất thông tin CCCD, có thể dùng OpenAI nếu được bật"""
    try:
        # Đọc text từ OCR cơ bản
        text_front = extract_text_with_ocr(image_front)
        text_back = extract_text_with_ocr(image_back)
        
        # Sử dụng OpenAI nếu được bật và có API key
        if use_openai and api_key and OPENAI_AVAILABLE:
            with st.spinner("🤖 Đang sử dụng OpenAI để trích xuất thông tin..."):
                openai_data = extract_cccd_with_openai(text_front, text_back, api_key)
                if openai_data:
                    st.success("✅ Đã sử dụng OpenAI để trích xuất thông tin")
                    return openai_data, text_front + "\n" + text_back
                else:
                    # Fallback về phương pháp cũ
                    st.info("ℹ️ Sử dụng phương pháp OCR thông thường")
                    info, full_text = extract_cccd_info(image_front, image_back)
                    return info, full_text
        else:
            info, full_text = extract_cccd_info(image_front, image_back)
            return info, full_text
            
    except Exception as e:
        st.error(f"Lỗi khi xử lý OCR: {str(e)}")
        return None, ""

def extract_cccd_info(image_front, image_back):
    """Trích xuất thông tin từ ảnh CCCD mặt trước và sau"""
    info = {
        'Số CCCD': '',
        'Họ và tên': '',
        'Ngày sinh': '',
        'Giới tính': '',
        'Quốc tịch': '',
        'Quê quán': '',
        'Nơi thường trú': '',
        'Ngày cấp': '',
        'Nơi cấp': ''
    }
    
    try:
        # Đọc text từ mặt trước
        text_front = extract_text_with_ocr(image_front)
        
        # Đọc text từ mặt sau
        text_back = extract_text_with_ocr(image_back)
        
        full_text = text_front + "\n" + text_back
        
        # Trích xuất số CCCD - định dạng "Số / No.: 080188012880"
        # Tìm từ khóa "Số / No.:" hoặc tương tự, sau đó lấy số 12 chữ số ngay sau đó
        so_no_pattern = r'(?:Số|SO)\s*[/\\]\s*No\.?\s*[:]'
        so_no_match = re.search(so_no_pattern, text_front, re.IGNORECASE)
        
        if so_no_match:
            # Lấy text sau "Số / No.:"
            text_after_label = text_front[so_no_match.end():]
            # Tìm số 12 chữ số đầu tiên ngay sau label (trong vòng 50 ký tự)
            number_match = re.search(r'\s*(\d{12})(?:\s|$|\n|[^\d])', text_after_label[:50])
            if number_match:
                cccd_number = number_match.group(1).replace(' ', '').replace('-', '').replace('.', '')
                if len(cccd_number) == 12 and cccd_number.isdigit():
                    info['Số CCCD'] = cccd_number
        else:
            # Fallback: thử các pattern khác nếu không tìm thấy "Số / No.:"
            cccd_patterns = [
                r'(?:Số|SO)\s*[/\\]?\s*No\.?\s*[:]?\s*(\d{12})(?:\s|$|\n)',  # "Số / No: 080188012880"
                r'(?:Số|SO)[\s:]*(\d{12})(?:\s|$|\n)',  # "Số: 080188012880"
                r'No\.\s*[:]?\s*(\d{12})(?:\s|$|\n)',  # "No.: 080188012880"
            ]
            for pattern in cccd_patterns:
                match = re.search(pattern, text_front, re.IGNORECASE | re.MULTILINE)
                if match:
                    cccd_number = match.group(1).replace(' ', '').replace('-', '').replace('.', '')
                    if len(cccd_number) == 12 and cccd_number.isdigit():
                        info['Số CCCD'] = cccd_number
                        break
        
        # Trích xuất họ và tên - tìm trong phạm vi rộng
        name_keyword_pattern = r'(?:Họ và tên|HỌ VÀ TÊN|Họ, chữ đệm và tên|Full name|Name)\s*[/\\]?\s*(?:Full name|Name)?\s*[:]'
        name_keyword_match = re.search(name_keyword_pattern, text_front, re.IGNORECASE)
        
        if name_keyword_match:
            # Lấy text trong phạm vi 150 ký tự sau từ khóa
            text_around_name = text_front[name_keyword_match.end():name_keyword_match.end() + 150]
            # Tìm tên (dòng chữ in hoa, có thể có nhiều từ)
            name_pattern = r'([A-ZÀ-Ỹ][A-ZÀ-Ỹ\s]{5,50}?)(?=\n|Ngày|Date|Giới|Sex|Gender|$)'
            match = re.search(name_pattern, text_around_name)
            if match:
                info['Họ và tên'] = match.group(1).strip()
        else:
            # Fallback: pattern thông thường
            name_patterns = [
                r'(?:Họ và tên|HỌ VÀ TÊN|Họ, chữ đệm và tên)[\s:]*([A-ZÀ-Ỹ][A-ZÀ-Ỹ\s]+?)(?:\n|Ngày)',
                r'(?:Full name|Name)[\s:]*([A-ZÀ-Ỹ][A-ZÀ-Ỹ\s]+?)(?:\n|Date)'
            ]
            for pattern in name_patterns:
                match = re.search(pattern, text_front, re.IGNORECASE)
                if match:
                    info['Họ và tên'] = match.group(1).strip()
                    break
        
        # Trích xuất ngày sinh - tìm trong phạm vi rộng quanh từ khóa
        dob_keyword_pattern = r'(?:Ngày sinh|Date of birth|DOB)\s*[/\\]?\s*Date of birth\s*[:]'
        dob_keyword_match = re.search(dob_keyword_pattern, text_front, re.IGNORECASE)
        
        if dob_keyword_match:
            # Lấy text trong phạm vi 100 ký tự sau từ khóa (để bắt ngày không thẳng hàng)
            text_around_dob = text_front[dob_keyword_match.start():dob_keyword_match.end() + 100]
            # Tìm ngày trong phạm vi này
            date_pattern = r'(\d{2})[\/\-](\d{2})[\/\-](\d{4})'
            match = re.search(date_pattern, text_around_dob)
            if match:
                day, month, year = match.groups()
                info['Ngày sinh'] = f"{day}/{month}/{year}"
        else:
            # Fallback: tìm pattern thông thường
            dob_patterns = [
                r'(?:Ngày sinh|Date of birth|DOB)[\s:/\\]*Date of birth\s*[:]\s*(\d{2})[\/\-](\d{2})[\/\-](\d{4})',
                r'(?:Ngày sinh|Date of birth|DOB)[\s:]*(\d{2})[\/\-](\d{2})[\/\-](\d{4})',
                r'(\d{2})[\/\-](\d{2})[\/\-](\d{4})'
            ]
            for pattern in dob_patterns:
                match = re.search(pattern, text_front, re.IGNORECASE | re.MULTILINE)
                if match:
                    day, month, year = match.groups()
                    info['Ngày sinh'] = f"{day}/{month}/{year}"
                    break
        
        # Trích xuất giới tính - tìm trong phạm vi rộng
        gender_keyword_pattern = r'(?:Giới tính|Sex|Gender)\s*[/\\]?\s*(?:Sex|Gender)?\s*[:]'
        gender_keyword_match = re.search(gender_keyword_pattern, text_front, re.IGNORECASE)
        
        if gender_keyword_match:
            # Lấy text trong phạm vi 50 ký tự sau từ khóa
            text_around_gender = text_front[gender_keyword_match.end():gender_keyword_match.end() + 50]
            gender_pattern = r'\s*((?:Nam|Nữ|Male|Female|NAM|NỮ))'
            match = re.search(gender_pattern, text_around_gender, re.IGNORECASE)
            if match:
                info['Giới tính'] = match.group(1).strip()
        else:
            # Fallback
            gender_patterns = [
                r'(?:Giới tính|Sex|Gender)[\s:]*((?:Nam|Nữ|Male|Female|NAM|NỮ))',
                r'(Nam|Nữ|Male|Female)'
            ]
            for pattern in gender_patterns:
                match = re.search(pattern, text_front, re.IGNORECASE)
                if match:
                    info['Giới tính'] = match.group(1).strip()
                    break
        
        # Trích xuất quốc tịch - tìm trong phạm vi rộng
        nationality_keyword_pattern = r'(?:Quốc tịch|Nationality)\s*[/\\]?\s*(?:Nationality)?\s*[:]'
        nationality_keyword_match = re.search(nationality_keyword_pattern, text_front, re.IGNORECASE)
        
        if nationality_keyword_match:
            # Lấy text trong phạm vi 100 ký tự sau từ khóa
            text_around_nationality = text_front[nationality_keyword_match.end():nationality_keyword_match.end() + 100]
            nationality_pattern = r'\s*([A-ZÀ-Ỹ\s]{2,50}?)(?=\n|Quê|Place|Origin|$)'
            match = re.search(nationality_pattern, text_around_nationality)
            if match:
                info['Quốc tịch'] = match.group(1).strip()
        else:
            # Fallback
            nationality_patterns = [
                r'(?:Quốc tịch|Nationality)[\s:]*([A-ZÀ-Ỹ\s]+?)(?:\n|Quê)',
                r'(Vietnam|Việt Nam|VN)'
            ]
            for pattern in nationality_patterns:
                match = re.search(pattern, text_front, re.IGNORECASE)
                if match:
                    info['Quốc tịch'] = match.group(1).strip() if match.lastindex and match.group(1) else "Việt Nam"
                    break
        
        # Trích xuất quê quán - thường ở dòng dưới, có thể nhiều dòng
        # Pattern linh hoạt hơn để tìm từ khóa
        que_quan_keyword_patterns = [
            r'Quê quán\s*[/\\]?\s*Place of origin\s*[:]',
            r'Quê quán\s*[:]',
            r'Place of origin\s*[:]'
        ]
        que_quan_keyword_match = None
        for pattern in que_quan_keyword_patterns:
            que_quan_keyword_match = re.search(pattern, text_front, re.IGNORECASE)
            if que_quan_keyword_match:
                break
        
        if que_quan_keyword_match:
            # Lấy text trong phạm vi 400 ký tự sau từ khóa
            text_after_keyword = text_front[que_quan_keyword_match.end():que_quan_keyword_match.end() + 400]
            
            # Tách thành các dòng (xử lý cả \n và \r\n)
            lines_after = re.split(r'\r?\n', text_after_keyword)
            
            # Thu thập các dòng địa chỉ (có thể nhiều dòng)
            address_lines = []
            
            # Đọc các dòng sau từ khóa (tối đa 4 dòng) cho đến khi gặp từ khóa mới
            for i, line in enumerate(lines_after[:4]):  # Xem 4 dòng đầu
                line = line.strip()
                # Loại bỏ từ khóa nếu còn sót
                line = re.sub(r'^(?:Quê quán|Place of origin|Origin)[\s:/\\]*', '', line, flags=re.IGNORECASE).strip()
                
                # Dừng nếu gặp từ khóa mới (Nơi thường trú, Permanent address, Quốc tịch)
                if re.match(r'^(?:Nơi thường trú|Permanent address|Address|Quốc tịch|Nationality)', line, re.IGNORECASE):
                    break
                # Thêm dòng nếu có vẻ là địa chỉ (bắt đầu bằng chữ hoa tiếng Việt, có dấu phẩy, hoặc có chữ cái)
                if line and (re.match(r'^[A-ZÀ-Ỹ]', line) or ',' in line):
                    address_lines.append(line)
                # Nếu dòng trống và đã có ít nhất 1 dòng địa chỉ, có thể đã kết thúc
                elif not line and address_lines:
                    break
            
            # Ghép các dòng lại thành địa chỉ đầy đủ
            if address_lines:
                info['Quê quán'] = ' '.join(address_lines).strip()
        
        # Fallback: pattern thông thường nếu chưa tìm được
        if not info.get('Quê quán'):
            que_quan_patterns = [
                r'Quê quán\s*[/\\]?\s*Place of origin\s*[:]\s*([A-ZÀ-Ỹ][A-ZÀ-Ỹ0-9/\s,\.\-]{5,150}?)(?=\n|Nơi|Permanent|Address|Quốc|Nationality|$)',
                r'Quê quán\s*[:]\s*([A-ZÀ-Ỹ][A-ZÀ-Ỹ0-9/\s,\.\-]{5,150}?)(?=\n|Nơi|$)',
                r'(?:Quê quán|Place of origin|Origin)[\s:]*([A-ZÀ-Ỹ0-9/\s,\.\-]{5,150}?)(?=\n|Nơi|Permanent|Address|Quốc|Nationality|$)'
            ]
            for pattern in que_quan_patterns:
                match = re.search(pattern, text_front, re.IGNORECASE | re.MULTILINE | re.DOTALL)
                if match:
                    value = match.group(1).strip()
                    if value:
                        info['Quê quán'] = value
                        break
        
        # Trích xuất nơi thường trú - có thể bắt đầu cùng dòng và tiếp tục ở dòng dưới
        search_text = text_back or text_front
        # Pattern linh hoạt hơn để tìm từ khóa
        thuong_tru_keyword_patterns = [
            r'Nơi thường trú\s*[/\\]?\s*Permanent address\s*[:]',
            r'Nơi thường trú\s*[/\\]?\s*Place of residence\s*[:]',
            r'Nơi thường trú\s*[:]',
            r'Permanent address\s*[:]'
        ]
        thuong_tru_keyword_match = None
        for pattern in thuong_tru_keyword_patterns:
            thuong_tru_keyword_match = re.search(pattern, search_text, re.IGNORECASE)
            if thuong_tru_keyword_match:
                break
        
        if thuong_tru_keyword_match:
            # Lấy text trong phạm vi 500 ký tự sau từ khóa (để bắt nhiều dòng)
            text_after_keyword = search_text[thuong_tru_keyword_match.end():thuong_tru_keyword_match.end() + 500]
            
            # Tách thành các dòng (xử lý cả \n và \r\n)
            lines_after = re.split(r'\r?\n', text_after_keyword)
            
            # Tìm phần còn lại trên cùng dòng (sau dấu :)
            first_line_after_colon = lines_after[0] if lines_after else ""
            # Loại bỏ từ khóa nếu còn sót
            first_line_after_colon = re.sub(r'^(?:Nơi thường trú|Permanent address|Address|Place of residence)[\s:/\\]*', '', first_line_after_colon, flags=re.IGNORECASE).strip()
            
            # Thu thập các dòng địa chỉ (có thể nhiều dòng)
            address_lines = []
            
            # Thêm phần còn lại trên dòng đầu nếu có (bắt đầu bằng số, chữ hoa, hoặc có dấu phẩy)
            if first_line_after_colon and (re.match(r'^[0-9A-ZÀ-Ỹ/]', first_line_after_colon) or ',' in first_line_after_colon or '.' in first_line_after_colon):
                address_lines.append(first_line_after_colon)
            
            # Đọc các dòng tiếp theo (tối đa 4 dòng) cho đến khi gặp từ khóa mới
            for i, line in enumerate(lines_after[1:5], start=1):  # Xem 4 dòng tiếp theo
                line = line.strip()
                # Dừng nếu gặp từ khóa mới (Ngày cấp, Date of issue, hoặc từ khóa khác)
                if re.match(r'^(?:Ngày cấp|Date of issue|Place of issue|Issued)', line, re.IGNORECASE):
                    break
                # Thêm dòng nếu có vẻ là địa chỉ (bắt đầu bằng số, chữ hoa, hoặc có dấu phẩy, dấu chấm)
                if line and (re.match(r'^[0-9A-ZÀ-Ỹ]', line) or ',' in line or '.' in line):
                    address_lines.append(line)
                # Nếu dòng trống và đã có ít nhất 1 dòng địa chỉ, có thể đã kết thúc
                elif not line and address_lines:
                    break
            
            # Ghép các dòng lại thành địa chỉ đầy đủ
            if address_lines:
                info['Nơi thường trú'] = ' '.join(address_lines).strip()
        
        # Fallback: pattern thông thường nếu chưa tìm được
        if not info.get('Nơi thường trú'):
            thuong_tru_patterns = [
                r'Nơi thường trú\s*[/\\]?\s*Permanent address\s*[:]\s*([0-9A-ZÀ-Ỹ/][A-ZÀ-Ỹ0-9/\s,\.\-]{10,200}?)(?=\n|Ngày|Date|$)',
                r'Nơi thường trú\s*[:]\s*([0-9A-ZÀ-Ỹ/][A-ZÀ-Ỹ0-9/\s,\.\-]{10,200}?)(?=\n|Ngày|$)',
                r'(?:Nơi thường trú|Permanent address|Place of residence)[\s:]*([0-9A-ZÀ-Ỹ/][A-ZÀ-Ỹ0-9/\s,\.\-]{10,200}?)(?=\n|Ngày|Date|$)'
            ]
            for pattern in thuong_tru_patterns:
                match = re.search(pattern, search_text, re.IGNORECASE | re.MULTILINE | re.DOTALL)
                if match:
                    value = match.group(1).strip()
                    if value:
                        info['Nơi thường trú'] = value
                        break
        
        # Trích xuất ngày cấp - tìm trong phạm vi rộng
        search_text_date = text_back or text_front
        # Pattern linh hoạt hơn để tìm từ khóa
        ngay_cap_keyword_patterns = [
            r'Ngày cấp\s*[/\\]?\s*Date of issue\s*[:]',
            r'Ngày cấp\s*[:]',
            r'Date of issue\s*[:]',
            r'Issued date\s*[:]'
        ]
        ngay_cap_keyword_match = None
        for pattern in ngay_cap_keyword_patterns:
            ngay_cap_keyword_match = re.search(pattern, search_text_date, re.IGNORECASE)
            if ngay_cap_keyword_match:
                break
        
        if ngay_cap_keyword_match:
            # Lấy text trong phạm vi 100 ký tự sau từ khóa
            text_around_ngay_cap = search_text_date[ngay_cap_keyword_match.end():ngay_cap_keyword_match.end() + 100]
            # Tìm ngày trong phạm vi này
            date_pattern = r'(\d{2})[\/\-](\d{2})[\/\-](\d{4})'
            match = re.search(date_pattern, text_around_ngay_cap)
            if match:
                day, month, year = match.groups()
                info['Ngày cấp'] = f"{day}/{month}/{year}"
        
        # Fallback: pattern thông thường nếu chưa tìm được
        if not info.get('Ngày cấp'):
            ngay_cap_patterns = [
                r'(?:Ngày cấp|Date of issue|Issued date)[\s:]*(\d{2})[\/\-](\d{2})[\/\-](\d{4})',
                r'(\d{2})[\/\-](\d{2})[\/\-](\d{4})'  # Tìm bất kỳ ngày nào trong text_back
            ]
            for pattern in ngay_cap_patterns:
                # Ưu tiên tìm trong text_back trước
                match = re.search(pattern, text_back or text_front, re.IGNORECASE)
                if match:
                    day, month, year = match.groups()
                    info['Ngày cấp'] = f"{day}/{month}/{year}"
                    break
        
        # Trích xuất nơi cấp
        noi_cap_patterns = [
            r'(?:Nơi cấp|Place of issue|Issued by)[\s:]*([A-ZÀ-Ỹ0-9/\s,]+?)(?:\n|$)',
            r'(?:Cơ quan cấp|Authority)[\s:]*([A-ZÀ-Ỹ0-9/\s,]+?)(?:\n|$)'
        ]
        for pattern in noi_cap_patterns:
            match = re.search(pattern, text_back or text_front, re.IGNORECASE | re.MULTILINE)
            if match:
                info['Nơi cấp'] = match.group(1).strip()
                break
        
        return info, full_text
        
    except Exception as e:
        st.error(f"Lỗi khi đọc OCR: {str(e)}")
        return info, ""

def load_excel_data():
    """Đọc dữ liệu từ file Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        # Lấy sheet đầu tiên (hoặc có thể chỉ định tên sheet)
        ws = wb.active
        
        # Lấy dữ liệu
        data = []
        headers = []
        
        # Đọc header từ hàng đầu tiên
        if ws.max_row > 0:
            for idx, cell in enumerate(ws[1]):
                header_value = cell.value if cell.value else ''
                # Xử lý cột trùng tên: thêm index cho cột trống
                if header_value == '':
                    header_value = f'Unnamed_{idx}'
                headers.append(header_value)
        
        # Đọc dữ liệu từ hàng 2 trở đi
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                data.append(row)
        
        if headers:
            # Xử lý cột trùng tên trước khi tạo DataFrame
            unique_headers = []
            header_counts = {}
            for header in headers:
                if header and not header.startswith('Unnamed_'):
                    # Đếm số lần xuất hiện của header này
                    count = header_counts.get(header, 0)
                    if count > 0:
                        unique_header = f"{header}_{count}"
                    else:
                        unique_header = header
                    header_counts[header] = count + 1
                    unique_headers.append(unique_header)
                else:
                    # Bỏ qua cột Unnamed
                    unique_headers.append(None)
            
            # Tạo DataFrame chỉ với các cột hợp lệ
            valid_indices = [i for i, h in enumerate(unique_headers) if h is not None]
            valid_headers = [unique_headers[i] for i in valid_indices]
            
            if data:
                valid_data = [[row[i] for i in valid_indices] for row in data]
                df = pd.DataFrame(valid_data, columns=valid_headers)
            else:
                df = pd.DataFrame(columns=valid_headers)
            
            return df
        else:
            return pd.DataFrame()
    except FileNotFoundError:
        st.warning(f"File {EXCEL_FILE} chưa tồn tại. Sẽ được tạo khi lưu dữ liệu đầu tiên.")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Lỗi khi đọc file Excel: {str(e)}")
        return pd.DataFrame()

def create_labor_contract(cccd_data, template_file="HDLD_Mau.txt"):
    """Tạo hợp đồng lao động từ template và dữ liệu CCCD"""
    try:
        # Đọc template
        with open(template_file, 'r', encoding='utf-8') as f:
            template = f.read()
        
        # Lấy ngày hiện tại
        today = datetime.now()
        current_date = today.strftime("%d/%m/%Y")
        current_day = today.strftime("%d")
        current_month = today.strftime("%m")
        current_year = today.strftime("%Y")
        
        # Thay thế các placeholder
        contract = template
        
        # Thông tin người lao động
        ho_ten = cccd_data.get('Họ và tên', '')
        ngay_sinh = cccd_data.get('Ngày sinh', '')
        gioi_tinh = cccd_data.get('Giới tính', '')
        quoc_tich = cccd_data.get('Quốc tịch', '')
        so_cccd = cccd_data.get('Số CCCD', '')
        ngay_cap = cccd_data.get('Ngày cấp', '')
        noi_cap = cccd_data.get('Nơi cấp', '')
        que_quan = cccd_data.get('Quê quán', '')
        thuong_tru = cccd_data.get('Nơi thường trú', '')
        
        # Xác định "Ông" hoặc "Bà" dựa vào giới tính
        xung_ho = "Ông/bà"
        if gioi_tinh and "Nam" in gioi_tinh:
            xung_ho = "Ông"
        elif gioi_tinh and "Nữ" in gioi_tinh:
            xung_ho = "Bà"
        
        # Thay thế các placeholder
        replacements = {
            '[Nguoi_LD]': ho_ten,
            '[Ngay_sinh]': ngay_sinh,
            '[Gioi_tinh]': gioi_tinh,
            '[Quoc_tich]': quoc_tich,
            '[So_CCCD]': so_cccd,
            '[Ngay_cap]': ngay_cap,
            '[Noi_cap]': noi_cap,
            '[Que_quan]': que_quan,
            '[DC_LH]': thuong_tru if thuong_tru else que_quan,
            'Ông/bà :': f'{xung_ho}:',
            'Hôm nay ngày ... tháng ... năm 2020': f'Hôm nay ngày {current_day} tháng {current_month} năm {current_year}',
            '...': 'Tp. Hồ Chí Minh',
        }
        
        for placeholder, value in replacements.items():
            contract = contract.replace(placeholder, value)
        
        return contract
    except Exception as e:
        st.error(f"Lỗi khi tạo hợp đồng: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

def generate_pdf_contract(contract_text, output_file):
    """Tạo file PDF từ nội dung hợp đồng với hỗ trợ tiếng Việt"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
        import os
        
        # Thử đăng ký font tiếng Việt
        font_name = 'Helvetica'  # Fallback
        try:
            # Thử Times New Roman (hỗ trợ tốt tiếng Việt)
            font_paths = [
                "C:/Windows/Fonts/times.ttf",
                "C:/Windows/Fonts/timesbd.ttf",  # Bold
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/arialbd.ttf",  # Bold
            ]
            
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        font_base_name = os.path.splitext(os.path.basename(font_path))[0]
                        pdfmetrics.registerFont(TTFont(font_base_name, font_path))
                        if 'times' in font_base_name.lower():
                            font_name = font_base_name
                            break
                    except:
                        continue
        except:
            pass
        
        # Tạo document
        doc = SimpleDocTemplate(output_file, pagesize=A4,
                               rightMargin=2*cm, leftMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        
        # Tạo style
        styles = getSampleStyleSheet()
        
        # Style cho đoạn văn thông thường
        normal_style = ParagraphStyle(
            'Normal_VN',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=11,
            leading=14,
            alignment=TA_LEFT,
            encoding='utf-8'
        )
        
        # Style cho tiêu đề (căn giữa)
        title_style = ParagraphStyle(
            'Title_VN',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=14,
            leading=18,
            alignment=TA_CENTER,
            encoding='utf-8'
        )
        
        # Tách nội dung thành các dòng và xử lý
        lines = contract_text.split('\n')
        story = []
        
        for line in lines:
            line = line.strip()
            if not line:
                story.append(Spacer(1, 0.2*cm))
                continue
            
            # Escape các ký tự đặc biệt cho HTML
            line_html = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            
            # Kiểm tra nếu là tiêu đề (chữ in hoa hoặc có định dạng đặc biệt)
            if line.isupper() and len(line) < 100 and any(keyword in line for keyword in ['CỘNG HÒA', 'HỢP ĐỒNG', 'NGƯỜI LAO ĐỘNG', 'NGƯỜI SỬ DỤNG']):
                para = Paragraph(line_html, title_style)
            else:
                para = Paragraph(line_html, normal_style)
            
            story.append(para)
            story.append(Spacer(1, 0.2*cm))
        
        # Build PDF
        doc.build(story)
        return True
    except Exception as e:
        st.error(f"Lỗi khi tạo PDF: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return False

def save_to_excel(new_data):
    """Ghi dữ liệu mới vào file Excel với định dạng font tiếng Việt và độ rộng cột"""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        
        headers = ['Số CCCD', 'Họ và tên', 'Ngày sinh', 'Giới tính', 'Quốc tịch', 
                  'Quê quán', 'Nơi thường trú', 'Ngày cấp', 'Nơi cấp']
        
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            
            # Kiểm tra xem đã có header chưa
            if ws.max_row == 0 or ws.cell(1, 1).value is None:
                ws.append(headers)
        except FileNotFoundError:
            # Tạo file mới nếu chưa tồn tại
            wb = openpyxl.Workbook()
            ws = wb.active
            # Thêm header
            ws.append(headers)
        
        # Định dạng header: font tiếng Việt, đậm, nền xanh
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Thêm dữ liệu mới
        row_data = [
            new_data.get('Số CCCD', ''),
            new_data.get('Họ và tên', ''),
            new_data.get('Ngày sinh', ''),
            new_data.get('Giới tính', ''),
            new_data.get('Quốc tịch', ''),
            new_data.get('Quê quán', ''),
            new_data.get('Nơi thường trú', ''),
            new_data.get('Ngày cấp', ''),
            new_data.get('Nơi cấp', '')
        ]
        ws.append(row_data)
        
        # Định dạng dữ liệu: font tiếng Việt, wrap text cho các cột dài
        data_font = Font(name="Arial", size=10)
        column_widths = {
            'A': 18,  # Số CCCD
            'B': 30,  # Họ và tên
            'C': 15,  # Ngày sinh
            'D': 12,  # Giới tính
            'E': 15,  # Quốc tịch
            'F': 50,  # Quê quán
            'G': 60,  # Nơi thường trú
            'H': 15,  # Ngày cấp
            'I': 50   # Nơi cấp
        }
        
        # Điều chỉnh độ rộng cột
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Định dạng dữ liệu cho hàng mới
        new_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(new_row, col_idx)
            cell.font = data_font
            # Wrap text cho các cột địa chỉ dài
            if col_idx in [6, 7, 9]:  # Quê quán, Nơi thường trú, Nơi cấp
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        st.error(f"Lỗi khi ghi file Excel: {str(e)}")
        import traceback
        st.error(f"Chi tiết lỗi: {traceback.format_exc()}")
        return False

# UI chính
tab1, tab2 = st.tabs(["📤 Nhập CCCD mới", "📋 Danh sách đã lưu"])

with tab1:
    st.header("Upload ảnh CCCD mặt trước và mặt sau")
    
    # Cấu hình OpenAI (nếu có)
    with st.expander("🔧 Cấu hình nâng cao (OpenAI API)", expanded=False):
        use_openai = st.checkbox("Sử dụng OpenAI API để trích xuất thông tin chính xác hơn", value=True)
        if use_openai:
            if not OPENAI_AVAILABLE:
                st.error("⚠️ Thư viện OpenAI chưa được cài đặt. Vui lòng chạy: pip install openai")
                api_key = None
            else:
                api_key = st.text_input(
                    "OpenAI API Key",
                    type="password",
                    help="Nhập API key của bạn từ https://platform.openai.com/api-keys",
                    value=st.session_state.get('openai_api_key', DEFAULT_API_KEY or '')
                )
                if api_key:
                    st.session_state['openai_api_key'] = api_key
                    st.success("✅ API Key đã được lưu")
        else:
            api_key = None
            use_openai = False
    
    # Khởi tạo biến nếu chưa có
    if 'use_openai' not in locals():
        use_openai = True if DEFAULT_API_KEY else False
    if 'api_key' not in locals():
        api_key = st.session_state.get('openai_api_key', DEFAULT_API_KEY)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Mặt trước")
        image_front_file = st.file_uploader(
            "Chọn ảnh mặt trước",
            type=['png', 'jpg', 'jpeg'],
            key="front"
        )
        if image_front_file:
            image_front = Image.open(image_front_file)
            st.image(image_front, caption="Mặt trước CCCD", use_container_width=True)
    
    with col2:
        st.subheader("Mặt sau")
        image_back_file = st.file_uploader(
            "Chọn ảnh mặt sau",
            type=['png', 'jpg', 'jpeg'],
            key="back"
        )
        if image_back_file:
            image_back = Image.open(image_back_file)
            st.image(image_back, caption="Mặt sau CCCD", use_container_width=True)
    
    if image_front_file and image_back_file:
        if st.button("🔍 Trích xuất thông tin", type="primary"):
            cccd_info, full_text = process_cccd_extraction(image_front, image_back, use_openai, api_key)
            
            # Đọc riêng OCR text từng mặt để hiển thị debug
            text_front_debug = extract_text_with_ocr(image_front)
            text_back_debug = extract_text_with_ocr(image_back)
            
            # Lưu vào session_state để giữ lại dữ liệu
            if cccd_info:
                st.session_state['cccd_info'] = cccd_info
                st.session_state['cccd_full_text'] = full_text
                st.session_state['text_front_debug'] = text_front_debug
                st.session_state['text_back_debug'] = text_back_debug
            
            # Hiển thị kết quả
            st.success("✅ Đã trích xuất thông tin!")
            
        # Hiển thị form chỉnh sửa thông tin (luôn hiển thị nếu có dữ liệu trong session_state)
        if 'cccd_info' in st.session_state and st.session_state['cccd_info']:
            cccd_info = st.session_state['cccd_info']
            
            # Hiển thị text OCR chi tiết (để debug) nếu có
            if 'text_front_debug' in st.session_state:
                with st.expander("🐛 DEBUG: Text OCR đã đọc (Để kiểm tra)", expanded=False):
                    st.write("**MẶT TRƯỚC (OCR Text):**")
                    st.text_area("", st.session_state['text_front_debug'], height=150, disabled=True, key="ocr_front_debug")
                    st.write("**MẶT SAU (OCR Text):**")
                    st.text_area("", st.session_state.get('text_back_debug', ''), height=150, disabled=True, key="ocr_back_debug")
                    st.write("**Full Text (Combined):**")
                    st.text_area("", st.session_state.get('cccd_full_text', ''), height=100, disabled=True, key="ocr_full_debug")
                    
                    # Kiểm tra xem có tìm thấy từ khóa không
                    st.write("**🔍 Kiểm tra từ khóa:**")
                    text_front = st.session_state.get('text_front_debug', '')
                    text_back = st.session_state.get('text_back_debug', '')
                    keywords_check = {
                        "Quê quán": bool(re.search(r'Quê quán|Place of origin', text_front, re.IGNORECASE)),
                        "Nơi thường trú": bool(re.search(r'Nơi thường trú|Permanent address|Place of residence', text_back or text_front, re.IGNORECASE)),
                        "Ngày cấp": bool(re.search(r'Ngày cấp|Date of issue', text_back or text_front, re.IGNORECASE)),
                    }
                    for key, found in keywords_check.items():
                        status = "✅ Tìm thấy" if found else "❌ KHÔNG tìm thấy"
                        st.write(f"- {key}: {status}")
            
            # Form chỉnh sửa thông tin
            st.markdown("### **Vui lòng kiểm tra và chỉnh sửa thông tin:**")
            
            col1, col2 = st.columns(2)
            
            with col1:
                so_cccd = st.text_input("Số CCCD", value=cccd_info.get('Số CCCD', ''))
                ho_ten = st.text_input("Họ và tên", value=cccd_info.get('Họ và tên', ''))
                ngay_sinh = st.text_input("Ngày sinh", value=cccd_info.get('Ngày sinh', ''))
                gioi_tinh = st.text_input("Giới tính", value=cccd_info.get('Giới tính', ''))
                quoc_tich = st.text_input("Quốc tịch", value=cccd_info.get('Quốc tịch', ''))
            
            with col2:
                que_quan = st.text_area("Quê quán", value=cccd_info.get('Quê quán', ''))
                thuong_tru = st.text_area("Nơi thường trú", value=cccd_info.get('Nơi thường trú', ''))
                ngay_cap = st.text_input("Ngày cấp", value=cccd_info.get('Ngày cấp', ''))
                noi_cap = st.text_input("Nơi cấp", value=cccd_info.get('Nơi cấp', ''))
            
            col_btn1, col_btn2 = st.columns(2)
            
            with col_btn1:
                if st.button("💾 Lưu vào Excel", type="primary", use_container_width=True):
                    final_data = {
                        'Số CCCD': so_cccd,
                        'Họ và tên': ho_ten,
                        'Ngày sinh': ngay_sinh,
                        'Giới tính': gioi_tinh,
                        'Quốc tịch': quoc_tich,
                        'Quê quán': que_quan,
                        'Nơi thường trú': thuong_tru,
                        'Ngày cấp': ngay_cap,
                        'Nơi cấp': noi_cap
                    }
                    
                    if save_to_excel(final_data):
                        st.success("✅ Đã lưu thông tin thành công vào file Excel!")
                        st.balloons()
                        # Xóa dữ liệu trong session_state sau khi lưu thành công
                        if 'cccd_info' in st.session_state:
                            del st.session_state['cccd_info']
                        if 'cccd_full_text' in st.session_state:
                            del st.session_state['cccd_full_text']
                        st.rerun()
                    else:
                        st.error("❌ Lỗi khi lưu thông tin")
            
            with col_btn2:
                if st.button("📄 Tạo hợp đồng lao động (PDF)", type="secondary", use_container_width=True):
                    final_data = {
                        'Số CCCD': so_cccd,
                        'Họ và tên': ho_ten,
                        'Ngày sinh': ngay_sinh,
                        'Giới tính': gioi_tinh,
                        'Quốc tịch': quoc_tich,
                        'Quê quán': que_quan,
                        'Nơi thường trú': thuong_tru,
                        'Ngày cấp': ngay_cap,
                        'Nơi cấp': noi_cap
                    }
                    
                    # Kiểm tra xem có đủ thông tin không
                    if not ho_ten or not so_cccd:
                        st.warning("⚠️ Vui lòng nhập đầy đủ thông tin (Họ và tên, Số CCCD) để tạo hợp đồng")
                    else:
                        with st.spinner("Đang tạo hợp đồng lao động..."):
                            # Tạo nội dung hợp đồng
                            contract_text = create_labor_contract(final_data)
                            
                            if contract_text:
                                # Tạo tên file PDF
                                safe_name = "".join(c for c in ho_ten if c.isalnum() or c in (' ', '-', '_')).strip()
                                pdf_filename = f"HDLD_{safe_name}_{so_cccd}.pdf"
                                
                                # Tạo file PDF
                                if generate_pdf_contract(contract_text, pdf_filename):
                                    st.success(f"✅ Đã tạo hợp đồng lao động: {pdf_filename}")
                                    
                                    # Đọc file PDF và cung cấp download
                                    with open(pdf_filename, "rb") as pdf_file:
                                        st.download_button(
                                            label="📥 Tải xuống hợp đồng (PDF)",
                                            data=pdf_file,
                                            file_name=pdf_filename,
                                            mime="application/pdf",
                                            type="primary"
                                        )
    
    elif image_front_file or image_back_file:
        st.warning("⚠️ Vui lòng upload cả 2 ảnh (mặt trước và mặt sau)")

with tab2:
    st.header("Danh sách thông tin đã lưu")
    
    df = load_excel_data()
    
    if not df.empty:
        st.dataframe(df, use_container_width=True)
        
        # Thống kê
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Tổng số bản ghi", len(df))
        with col2:
            if st.button("🔄 Làm mới dữ liệu"):
                st.rerun()
    else:
        st.info("Chưa có thông tin nào được lưu. Vui lòng nhập CCCD mới ở tab 'Nhập CCCD mới'")

st.markdown("---")
st.markdown(f"**File Excel:** `{EXCEL_FILE}`")
